import numpy as np
from scipy.optimize import fsolve
from scipy.integrate import quad
import math
from scipy.optimize import root
import csv
from scipy.optimize import minimize
from tkinter import Entry, Button, Label, Tk
import openpyxl
###################
# constants
################
m = 9.10938356E-31  # mass of electron in Kg
kb = 1.3806505E-23  # boltzmann constant (J/K)
kb_eV = 8.6173E-5  # boltzmann constant (eV/K)
e = 1.60217653E-19  # Charge on electron (Coulomb)
R = 8.314  # J / mol K (This is the gas constant)
h = 6.62607015e-34  # This is plancks contant in J * s
h_eV = 4.135667696E-15  # Plancks constant in eV * s
integralLowerLimit = 0  # Set the lower limit of the fermi integrals
integralUpperLimit = 100  # Set the upper limit of the fermi integrals
hbar = h/(2*math.pi)
pi = math.pi

# Note on scattering parameter, r:
# in this code, an r = 0 corresponds with acoustic phonon scattering

file_path = 'Terra_SnTe_roomtemp.xlsx'
workbook = openpyxl.load_workbook(file_path)

seebeck_sheet = workbook['Seebeck (uVK)']
cc_sheet = workbook['Hall carrier (cm^-3)']
resistive_sheet = workbook['resistivity (mOhm-cm)']
nernst_sheet = workbook['Nernst (uVKT)']


if 'MOFC' not in workbook.sheetnames:
    mosc_sheet = workbook.create_sheet(title='MOFC')
else:
    mosc_sheet = workbook['MOFC']

headers = ['Reduced Fermi level',
                    'm*/me',
                    'Tau_0 (*1E-15 sec)',
                    'r',
                    'Calculated loss']
mosc_sheet.append(headers)

#find # of rows in sheet
num_rows = seebeck_sheet.max_row
T = 273 + 50

def FermiIntegrand(epsilon, eta, j): #epsilon is reduced energy and your integration variable, eta is reduced fermi level, j is order of integral; j = 0 is just integral of FD dist
    return epsilon ** j / (1 + np.exp(epsilon - eta))


def directsolveSeebeck(eta, r): #expected 5.7880022 for seebeck 48.6, r=0
    seebeck = kb/e * ((r+2)*quad(FermiIntegrand, 0, 10, args=(eta, r+1))[0]/((r+1)*(quad(FermiIntegrand, 0, 10, args=(eta, r))[0])) - eta) # 0 = seeb - expression with Fermi integrals
    return seebeck * 1E6


def directsolveHallCoefficient(eta, effmass, T, r): #expected: 1.2275853 for print(fsolve(nowweEffectivemass, 1, args=(3.87E+20, 273+40), full_output=1)) and args = (5.7880022) on integrals
    R_H = (1/e)*(((2*r+0.5) * quad(FermiIntegrand, 0, 100, args=(eta, 2*r-0.5))[0])/
                                  ((((r+1)*quad(FermiIntegrand, 0, 100, args=(eta,r))[0]))**2)*((3*pi**2*hbar**3)/((2*effmass*m*kb*T)**(3/2))))
    return 1E-6/(e*R_H)


def directsolveconductivity(tau_0, effmass, T, eta, r):
    effmass = effmass*m
    tau_0 = tau_0 * 1E-15
    conductivity = (((e**2)*(2*kb*T)**(3/2))/(3*pi**2*hbar**3))*effmass**0.5*tau_0*(r+1)*quad(FermiIntegrand, 0, 100, args=(eta,r))[0]
    return 1E5/conductivity


def directsolveNernst(tau_0, effmass, eta, r):
    effmass = effmass*m
    tau_0 = tau_0 * 1E-15
    nernst = kb*tau_0/effmass * ((r+1) * quad(FermiIntegrand, 0, 100, args= (eta, r))[0] * (2*r + 3/2) * quad(FermiIntegrand, 0, 100, args= (eta, 2*r+1/2))[0] - (r + 2)*(2*r + 1/2) *
    quad(FermiIntegrand, 0, 100, args= (eta, r+1))[0] * quad(FermiIntegrand, 0, 100, args=(eta,2*r-1/2))[0]) / (((r+1)*quad(FermiIntegrand,0,100, args=(eta, r))[0])**2)
    return nernst*1E6


def objective_fcn(x): #this is the loss function we want to minimize
    eta = x[0]
    effmass = x[1]
    tau_0 = x[2]
    r = x[3]
    seebeck_calc = directsolveSeebeck(eta,r)
    cc_calc = directsolveHallCoefficient(eta, effmass, T, r)
    rho_calc = directsolveconductivity(tau_0, effmass, T, eta, r)
    nernst_calc = directsolveNernst(tau_0, effmass,eta,r)

    ### objectivefunction below
    return ((seebeck_truth - seebeck_calc)/seebeck_truth) ** 2 +  \
            ((cc_truth - cc_calc)/cc_truth) ** 2 +       \
            ((rho_truth - rho_calc)/rho_truth) ** 2 +   \
            ((nernst_truth - nernst_calc)/nernst_truth) ** 2

bounds_eta = (-5,35) #positive means into the band, negative means into the band gap
bounds_effmass = (0.001,5) #these are in units of m_e, i,e, m*/m_e = eff_mass
bounds_tau_0 = (.001,1000) #these are in units of femtoseconds (E-15)
bounds_r = (0,2) #0 means acoustic phonon scattering, 2 = ionized impurity scattering

bounds = [bounds_eta, bounds_effmass, bounds_tau_0, bounds_r]

#seed values
x0 = [1, 1, 1, 1]


# Iterate over rows, process inputs, and write results
for row_idx in range(2, num_rows+1):  # Adjust row range as needed
    # Get values from corresponding sheets
    seebeck_truth = float(seebeck_sheet.cell(row=row_idx, column=3).value)
    print(seebeck_truth)
    cc_truth = float(cc_sheet.cell(row=row_idx, column=3).value)
    rho_truth = float(resistive_sheet.cell(row=row_idx, column=3).value)
    print(rho_truth)
    nernst_truth = float(nernst_sheet.cell(row=row_idx, column=3).value)

    # Ensure inputs are valid (skip row if any input is missing)
    if None in [seebeck_truth, cc_truth, rho_truth, nernst_truth]:
        continue
    # seed values
    x0 = [1, 1, 1, 1]
    result = minimize(objective_fcn, x0, method='trust-constr', bounds=bounds)
    # Get function results and write them to MOFC sheet, or whatever you named it in line 37
    resultslist = []
    for item in result.x:
        resultslist.append(item)
    resultslist.append(result.fun)
    mosc_sheet.append(resultslist)
    workbook.save(file_path)

# Save the updated workbook
workbook.save(file_path)
